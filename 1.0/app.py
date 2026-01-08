from flask import Flask, render_template, request, redirect, url_for, send_file, session
from werkzeug.utils import secure_filename
from pathlib import Path
import io
import csv
import zipfile
import markdown

import sae105_programme as core  # ton script d'analyse

app = Flask(__name__)
app.secret_key = "change_me"

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Nettoyage au démarrage : on supprime les anciens fichiers du dossier uploads
for f in UPLOAD_DIR.iterdir():
    if f.is_file():
        try:
            f.unlink()
        except OSError:
            pass

# ------------------ langue ------------------ #

@app.before_request
def set_lang():
    lang = request.args.get("lang")
    if lang in ("fr", "en"):
        session["lang"] = lang
    if "lang" not in session:
        session["lang"] = "fr"
    core.current_lang = session["lang"]

# ------------------ routes ------------------ #

@app.route("/")
def index():
    # valeurs par défaut = celles de ton script
    return render_template(
        "index.html",
        lang=session.get("lang", "fr"),
        default_scan_ports=core.DEFAULT_SCAN_PORTS,
        default_dos_abs=core.DEFAULT_DOS_ABS,
        default_dos_pct=core.DEFAULT_DOS_PCT,
        default_noisy_dests=core.DEFAULT_NOISY_DESTS,
        default_noisy_pkts=core.DEFAULT_NOISY_PKTS,
        default_syn_abs=core.DEFAULT_SYN_ABS,
        default_syn_ratio=core.DEFAULT_SYN_RATIO,
    )

@app.route("/analyser", methods=["POST"])
def analyser():
    f = request.files.get("file")
    if not f or f.filename == "":
        return redirect(url_for("index"))

    # Supprimer le précédent fichier analysé si présent
    last = session.get("last_file")
    if last:
        p_last = Path(last)
        if p_last.exists():
            try:
                p_last.unlink()
            except OSError:
                pass

    filename = secure_filename(f.filename)
    fpath = UPLOAD_DIR / filename
    f.save(fpath)

    # 1) parse tcpdump
    packets, first_ts, last_ts = core.parse_tcpdump_file(fpath)
    if not packets:
        return redirect(url_for("index"))

    # 2) récupérer les seuils envoyés par le formulaire
    def get_int(name, default):
        v = request.form.get(name, "").strip()
        try:
            return int(v) if v else default
        except ValueError:
            return default

    def get_float(name, default):
        v = request.form.get(name, "").strip()
        try:
            return float(v) if v else default
        except ValueError:
            return default

    th_scan_ports = get_int("scan_ports", core.DEFAULT_SCAN_PORTS)
    th_dos_abs = get_int("dos_abs", core.DEFAULT_DOS_ABS)
    th_dos_pct = get_float("dos_pct", core.DEFAULT_DOS_PCT)  # % entier
    th_noisy_dests = get_int("noisy_dests", core.DEFAULT_NOISY_DESTS)
    th_noisy_pkts = get_int("noisy_pkts", core.DEFAULT_NOISY_PKTS)
    th_syn_abs = get_int("syn_abs", core.DEFAULT_SYN_ABS)
    th_syn_ratio = get_float("syn_ratio", core.DEFAULT_SYN_RATIO)  # %

    # 3) stats
    (
        by_src,
        by_dst,
        by_dport,
        by_flow,
        flows_per_src,
        total_bytes,
        flags_counter,
        syn_per_dst,
        syn_total,
        by_proto,
        bytes_per_src,
    ) = core.compute_stats(packets)

    # 4) détection avec les seuils fournis
    scans = core.detect_port_scans(by_flow, threshold_ports=th_scan_ports)
    dos = core.detect_dos(
        by_dst,
        abs_threshold=th_dos_abs,
        pct_threshold=th_dos_pct / 100.0,
    )
    noisy = core.detect_noisy_sources(
        by_src,
        flows_per_src,
        dest_threshold=th_noisy_dests,
        pkt_threshold=th_noisy_pkts,
    )
    synfloods = core.detect_syn_flood(
        syn_per_dst,
        by_dst,
        syn_abs=th_syn_abs,
        syn_ratio=th_syn_ratio / 100.0,
    )
    alerts = scans + dos + noisy + synfloods

    # 5) graphes
    charts = core.build_chart_images(by_src, by_dst, by_dport, bytes_per_src)

    # 6) rapport complet Markdown -> HTML
    markdown_report = core.build_markdown_report(
        packets,
        by_src,
        by_dst,
        by_dport,
        by_flow,
        alerts,
        first_ts,
        last_ts,
        total_bytes,
        flags_counter,
        syn_per_dst,
        syn_total,
        by_proto,
    )
    html_report = markdown.markdown(markdown_report, extensions=["tables"])

    # stocker pour download CSV/Excel
    session["last_file"] = str(fpath)

    return render_template(
        "rapport.html",
        filename=filename,
        charts=charts,
        html_report=html_report,
        scan_ports=th_scan_ports,
        dos_abs=th_dos_abs,
        dos_pct=th_dos_pct,
        noisy_dests=th_noisy_dests,
        noisy_pkts=th_noisy_pkts,
        syn_abs=th_syn_abs,
        syn_ratio=th_syn_ratio,
        total_bytes=total_bytes,
    )

# ------------------ téléchargements ------------------ #

@app.route("/download/csv")
def download_csv():
    last = session.get("last_file")
    if not last:
        return redirect(url_for("index"))
    path = Path(last)
    if not path.exists():
        return redirect(url_for("index"))

    packets, first_ts, last_ts = core.parse_tcpdump_file(path)
    (
        by_src,
        by_dst,
        by_dport,
        by_flow,
        flows_per_src,
        total_bytes,
        flags_counter,
        syn_per_dst,
        syn_total,
        by_proto,
        bytes_per_src,
    ) = core.compute_stats(packets)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        raw_csv = io.StringIO()
        writer = csv.DictWriter(raw_csv, fieldnames=list(packets[0].keys()))
        writer.writeheader()
        writer.writerows(packets)
        z.writestr("packets_raw.csv", raw_csv.getvalue())

        src_csv = io.StringIO()
        w = csv.writer(src_csv)
        w.writerow(["src_host", "packets"])
        for host, count in by_src.most_common():
            w.writerow([host, count])
        z.writestr("stats_sources.csv", src_csv.getvalue())

        dst_csv = io.StringIO()
        w = csv.writer(dst_csv)
        w.writerow(["dst_host", "packets"])
        for host, count in by_dst.most_common():
            w.writerow([host, count])
        z.writestr("stats_destinations.csv", dst_csv.getvalue())

        port_csv = io.StringIO()
        w = csv.writer(port_csv)
        w.writerow(["dst_port", "packets"])
        for port, count in by_dport.most_common():
            w.writerow([port, count])
        z.writestr("stats_ports.csv", port_csv.getvalue())

        syn_csv = io.StringIO()
        w = csv.writer(syn_csv)
        w.writerow(["dst_host", "syn_packets"])
        for dst, count in syn_per_dst.most_common():
            w.writerow([dst, count])
        z.writestr("stats_syn_per_dst.csv", syn_csv.getvalue())

        proto_csv = io.StringIO()
        w = csv.writer(proto_csv)
        w.writerow(["protocol", "packets"])
        for proto, count in by_proto.most_common():
            w.writerow([proto, count])
        z.writestr("stats_protocols.csv", proto_csv.getvalue())

    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="tcpdump_stats_csv.zip")

@app.route("/download/xlsx")
def download_xlsx():
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, BarChart, Reference

    last = session.get("last_file")
    if not last:
        return redirect(url_for("index"))
    path = Path(last)
    if not path.exists():
        return redirect(url_for("index"))

    packets, first_ts, last_ts = core.parse_tcpdump_file(path)
    (
        by_src,
        by_dst,
        by_dport,
        by_flow,
        flows_per_src,
        total_bytes,
        flags_counter,
        syn_per_dst,
        syn_total,
        by_proto,
        bytes_per_src,
    ) = core.compute_stats(packets)

    wb = Workbook()
    wb.remove(wb.active)

    wsraw = wb.create_sheet("Raw")
    headers = list(packets[0].keys())
    wsraw.append(headers)
    for p in packets:
        wsraw.append([p.get(h) for h in headers])

    wssrc = wb.create_sheet("Sources")
    wssrc.append(["src_host", "packets"])
    for host, count in by_src.most_common():
        wssrc.append([host, count])
    piesrc = PieChart()
    piesrc.title = "Traffic by source"
    labels = Reference(wssrc, min_col=1, min_row=2, max_row=wssrc.max_row)
    data = Reference(wssrc, min_col=2, min_row=1, max_row=wssrc.max_row)
    piesrc.add_data(data, titles_from_data=True)
    piesrc.set_categories(labels)
    wssrc.add_chart(piesrc, "E2")

    wsdst = wb.create_sheet("Destinations")
    wsdst.append(["dst_host", "packets"])
    for host, count in by_dst.most_common():
        wsdst.append([host, count])
    piedst = PieChart()
    piedst.title = "Traffic by destination"
    labels = Reference(wsdst, min_col=1, min_row=2, max_row=wsdst.max_row)
    data = Reference(wsdst, min_col=2, min_row=1, max_row=wsdst.max_row)
    piedst.add_data(data, titles_from_data=True)
    piedst.set_categories(labels)
    wsdst.add_chart(piedst, "E2")

    wsport = wb.create_sheet("Ports")
    wsport.append(["dst_port", "packets"])
    for port, count in by_dport.most_common():
        wsport.append([str(port), count])
    barport = BarChart()
    barport.title = "Most used destination ports"
    labels = Reference(wsport, min_col=1, min_row=2, max_row=wsport.max_row)
    data = Reference(wsport, min_col=2, min_row=1, max_row=wsport.max_row)
    barport.add_data(data, titles_from_data=True)
    barport.set_categories(labels)
    barport.y_axis.title = "Packets"
    barport.x_axis.title = "Port"
    wsport.add_chart(barport, "E2")

    wssyn = wb.create_sheet("SYN")
    wssyn.append(["dst_host", "syn_packets"])
    for dst, count in syn_per_dst.most_common():
        wssyn.append([dst, count])
    barsyn = BarChart()
    barsyn.title = "SYN packets by destination"
    labels = Reference(wssyn, min_col=1, min_row=2, max_row=wssyn.max_row)
    data = Reference(wssyn, min_col=2, min_row=1, max_row=wssyn.max_row)
    barsyn.add_data(data, titles_from_data=True)
    barsyn.set_categories(labels)
    barsyn.y_axis.title = "SYN packets"
    wssyn.add_chart(barsyn, "E2")

    wssum = wb.create_sheet("Résumé")
    wssum.append(["Indicateur", "Valeur"])
    wssum.append(["Paquets analysés", len(packets)])
    wssum.append(["Volume total (octets)", total_bytes])
    wssum.append(["Sources distinctes", len(by_src)])
    wssum.append(["Destinations distinctes", len(by_dst)])
    wssum.append(["Ports distincts", len(by_dport)])
    wssum.append(["Protocoles distincts", len(by_proto)])
    wssum.append(["Paquets SYN totaux", syn_total])

    # Onglet volume par source
    wsvol = wb.create_sheet("Volume_Sources")
    wsvol.append(["src_host", "volume_octets"])
    for host, vol in bytes_per_src.most_common():
        wsvol.append([host, vol])

    barvol = BarChart()
    barvol.title = "Top sources par volume"
    labels = Reference(wsvol, min_col=1, min_row=2, max_row=wsvol.max_row)
    data = Reference(wsvol, min_col=2, min_row=1, max_row=wsvol.max_row)
    barvol.add_data(data, titles_from_data=True)
    barvol.set_categories(labels)
    barvol.y_axis.title = "Octets"
    barvol.x_axis.title = "Source"
    wsvol.add_chart(barvol, "E2")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="rapport_tcpdump.xlsx")

if __name__ == "__main__":
    import threading
    import time
    import webbrowser

    def open_browser():
        # petit délai pour laisser le serveur démarrer
        time.sleep(1)
        webbrowser.open("http://127.0.0.1:5000/")

    threading.Thread(target=open_browser, daemon=True).start()
    app.run(debug=False)
